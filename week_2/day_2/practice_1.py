# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module
from openpyxl import load_workbook
# Import the workbook that you created in yesterday's exercise from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook('/Users/ganegodagerodrigo/Desktop/VetsinTech/python3-course/week_2/day_1/inventory.xlsx')

# verify what sheet names are available
print(wb.sheetnames)
# access and store the appropriate worksheet to the variable 'ws' 
ws = wb['CURRENT_MONTH_INVENTORY']

# Print out the cell values for each row
all_rows = ws.rows

for row in all_rows:
    for each_cell in row:
        print(each_cell.value)

# Create a new column within that worksheet called order_amount

# save the latest changes
wb.save("./inventory.xlsx")