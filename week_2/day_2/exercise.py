# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module
from openpyxl import load_workbook
# Import the workbook that you updated in today's practice from
wb = load_workbook('./inventory.xlsx')
# "./spreadsheets/inventory.xlsx"

# access and store the appropriate worksheet to the variable 'ws'
ws = wb.active
# print(ws.title)
# Define a function called add_order_amount that takes in a single parameter called 'row'
# def add_order_amount(row):
def add_order_amount(row):
    
    # IF the quantity is less or equal to than the threshold,
        max_amount = int(ws.cell(row=row, column=3).value)
        quantity = int(ws.cell(row=row, column=5).value)
        threshold = int(ws.cell(row=row, column=4).value)

        if quantity <= threshold:

        # than calculate the order_amount (max_amount - quantity) 
            order_amount = max_amount - quantity
        # assign the value to that row, column 6
            ws.cell(row=row, column=6,value = order_amount)
# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

# Perform a for..in loop through the range(2, len(inventory.rows))
#   - call the function add_order_amount() passing in the number of the range
for row in range(2, ws.max_row): # max_row counts the number of rows and starts at 1. it is not inclusive. it is the last row + 1
    add_order_amount(row)



wb.save("./inventory.xlsx2")