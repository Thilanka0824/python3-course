# Magic 8 ball!

import random
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

fortunes = ['It is certain',
              'It is decidedly so',
              'Yes',
              'Reply hazy try again',
              'Ask again later',
              'Concentrate',
              'My reply is no',
              'Outlook not so good',
              'Very doubtful']


def get_fortune():
    return random.choice(fortunes)

fate = get_fortune()

print(fate)

# Given the function above, it will randomly choose from the fortunes list and print one.
# Using this concept we will create our own function for a new workbook in openpyxl

# Create a list called first_names and prepopulate with 10 custom first names
first_names = ['John', 'Jane', 'Jack', 'Jill', 'James', 'Jenny', 'Jared', 'Jasmine', 'Jesse', 'Jocelyn']
# Create a list called last_names and prepopulate with 10 custom last names
last_names = ['Smith', 'Doe', 'Kohnson', 'Vackson', 'Wones', 'Lenkins', 'Zefferson', 'Mennings', 'Wensen', 'Pennings']
# Define a function called assign_names with a parameter 'row'
#   - cell at row=row and column=1 assign the value to str(random.randint(111111, 999999))
#   - cell at row=row and column=2 assign the value to a random choice of first name
#   - cell at row=row and column=3 assign the value to a random choice of last name
def assign_names(row):
    ws.cell(row=row, column=1, value = str(random.randint(111111, 999999)))
    ws.cell(row=row, column=2, value = random.choice(first_names))
    ws.cell(row=row, column=3, value = random.choice(last_names))


# set up appropriately for a new workbook and worksheet

# loop through the range of 1-10 and for each number in the range
# - call/invoke the assign_names function while passing in the number as the 'row' argument
for cell in range(1, 11):
    assign_names(cell)

# save your new workbook!
wb.save("../spreadsheets/day_3_exercise.xlsx")