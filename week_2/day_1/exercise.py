# ITP Week 2 Day 1 Exercise

# SCENARIO: You're a store owner receiving the inventory report for this month.
# You will receive the product order spreadsheet soon and it is easier to calculate your order
# if your inventory was also on a spreadsheet! Recreate the following spreadsheet with Python:
from openpyxl import Workbook

# created workbook
wb = Workbook()

# active sheet
CURRENT_MONTH_INVENTORY = wb.active 

CURRENT_MONTH_INVENTORY["A1"] = "Product Name"
CURRENT_MONTH_INVENTORY["B1"] = "Product ID"
CURRENT_MONTH_INVENTORY["C1"] = "Max Amount"
CURRENT_MONTH_INVENTORY["D1"] = "Reorder Threshold"
CURRENT_MONTH_INVENTORY["E1"] = "Quantity"

productNames = [
    "oreo",
    "coke",
    "pepsi",
    "lays_chip",
    "pringles",
    "sour_worms",
    "choco_cookies",
    "donuts",
    "hot_dogs",
    "ice_cream",
    "gum",
    "pretzels",
    "kit_kat",
]

productID = [
    2323,
    6545,
    3456,
    4567,
    2134,
    2362,
    923,
    2786,
    6723,
    9237,
    2092,
    8246,
    9276,
]

maxAmount = [1000, 500, 200, 1500, 2000, 100, 200, 200, 100, 200, 3500, 100, 1000]

reorderThreshold = [300, 100, 50, 500, 600, 10, 25, 25, 10, 50, 1000, 5, 250]

quantity = [743, 101, 137, 364, 120, 85, 24, 12, 39, 234, 1232, 11, 249]

dataNames = [productNames, productID, maxAmount, reorderThreshold, quantity]
# don't forget your appropriate imports.
# assign the title of the initial active sheet to "CURRENT_MONTH_INVENTORY"


listLength = len(productNames)
for cell in range(2, listLength + 1):
    CURRENT_MONTH_INVENTORY.cell(row=cell, column=1, value=productNames[cell - 2])
    CURRENT_MONTH_INVENTORY.cell(row=cell, column=2, value=productID[cell - 2])
    CURRENT_MONTH_INVENTORY.cell(row=cell, column=3, value=maxAmount[cell - 2])
    CURRENT_MONTH_INVENTORY.cell(row=cell, column=4, value=reorderThreshold[cell - 2])
    CURRENT_MONTH_INVENTORY.cell(row=cell, column=5, value=quantity[cell - 2])


#   product_name    product_id  max_amount      reorder_threshold   quantity
# 	oreo            2323        1000            300                 743
# 	coke            6545        500             100                 101
# 	pepsi	        3456        200             50                  137
# 	lays_chip	    4567        1500            500                 364
# 	pringles	    2134        2000            600                 120
# 	sour_worms	    2362        100             10                  85
# 	choco_cookies   0923	    200             25                  24
# 	donuts	        2786        200             25                  12
# 	hot_dogs	    6723        100             10                  39
# 	ice_cream	    9237        200             50                  234
# 	gum	            2092        3500            1000                1232
# 	pretzels        8246	    100             5                   11
# 	kit_kat	        9276        1000            250                 249


# TIP: create a list of each column (ie. product_names = [oreo, ...]) and use those to loop through :)

# save your file
wb.save("./inventory2.xlsx")
