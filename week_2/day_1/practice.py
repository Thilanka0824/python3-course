# ITP Week 2 Day 1 (In-Class) Practice

# A1. from the appropriate library, import only the Workbook
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with..
wb = Workbook()
print(type(wb))
print(wb.active)
# A3. We need to interact with a single worksheet.
ws = wb.active
# A4. assign the value of "First Name" to A1
ws['A1'] = "First Name"
# A5. assign the value of "Last Name" to B1
ws['B1'] = "Last Name"
# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)
for cell in range(2, 11):
    # ws['A' + str(cell)] = "Gabriel"
    print(ws.cell(row=cell, column=1, value = "Gabriel")) 


# for row in ws.iter_rows(min_row=1, max_row=10, min_col=5, max_col=1):
#     for cell in row:
#         print(cell.value = "Gabriel")



last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list
for num in range(2, 11):
    ws.cell(row=num, column=2, value=last_names[num-2])

# B3. Save the file
wb.save("./day_1_lecture.xlsx")


