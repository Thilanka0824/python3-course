from openpyxl import Workbook
wb = Workbook()
ws = wb.active

item_oreo = {
    "product_id": 2323,
    "name": "oreo",
    "reorder_threshold": 300,
    "inventory": 743,
    "max_amount": 1000,
    "description": "yummy yummy"
}

item_coke = {
    "product_id": 6545,
    "name": "coke",
    "reorder_threshold": 100,
    "inventory": 101,
    "max_amount": 500
}

item_pepsi = {
    "product_id": 3456,
    "name": "pepsi",
    "reorder_threshold": 50,
    "inventory": 137,
    "max_amount": 200
}

# print(ws.cell(row=1, column=1, value=item_oreo['product_id']))
# print(ws.cell(row=1, column=2, value=item_oreo['name']))
# print(ws.cell(row=1, column=3, value=item_oreo['reorder_threshold']))
# print(ws.cell(row=1, column=4, value=item_oreo['inventory']))
# print(ws.cell(row=1, column=5, value=item_oreo['max_amount']))

def dict_to_row(item):
    print(ws.cell(row=1, column=1, value=item_oreo['product_id']))
    print(ws.cell(row=1, column=2, value=item_oreo['name']))
    print(ws.cell(row=1, column=3, value=item_oreo['reorder_threshold']))
    print(ws.cell(row=1, column=4, value=item_oreo['inventory']))
    print(ws.cell(row=1, column=5, value=item_oreo['max_amount']))

# print(len(item_oreo.keys()))

for number in range(1, len(item_oreo.keys()) + 1):
    ws.cell(row=1, column=number, value=list(item_oreo.values())[number - 1])



# column_count = 1
# for key in item_oreo:
#     print(ws.cell(row=1, column=column_count, value=item_oreo[key]).value)
#     column_count += 1

another_col_count = 1

for key, value in item_oreo.items():
    ws.cell(row=1, column=another_col_count, value=value)
    another_col_count+=1


# wb.save("./demo.xlsx")

#enumerate

# list =[1, 2, 3]

# print(enumerate([1,2,3]))

# for count, value in enumerate(item_oreo.values(), 2):
#     ws.cell(row=1, column=count, value=value)

inventory_list =[item_oreo, item_coke, item_pepsi]

for row_count, each_dict in enumerate(inventory_list, 1):
    for col_count, value in enumerate(each_dict.values(), 1):
        ws.cell(row=row_count, column=col_count, value=value)

# wb.save("./demo3.xlsx")

store = {
    "aisle_1": ["fruits", "vegetables", "dairy"],
    "aisle_2": ["meat", "frozen", "bread"],
    "aisle_3": ["poultry", "seafood", "dressing"],
    "aisle_4": ["cereal", "soup", "water"],
}

small_store = {
    "single_aisle": [item_coke, item_oreo, item_pepsi]
}

print(small_store["single_aisle"][2]["product_id"])


complex_store = {
    "one_aisle": {
        "dairy": [{"name": "milk", "price": 2.59}, {"name": "cheese", "price": 1.59}],
        "bread": [{"name": "wonderbread", "price": 2.59}, {"name": "wheat", "price": 1.59}]
    },
    "two_aisle": {
        "meat": [{"name": "beef", "price": 6.59}, {"name": "pork", "price": 7.59}]
    },
}

print(complex_store["one_aisle"]['dairy'][0]['price'])
print(complex_store["two_aisle"]['meat'][1])

# wb.save("./demo4.xlsx")