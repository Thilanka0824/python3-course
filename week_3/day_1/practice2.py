# ITP Week 3 Day 1 Practice

# import your required modules/methods
from openpyxl import Workbook()

wb = Workbook()

ws = wb.active


# given the following items, using the methods we covered, write to openpyxl


# use an external counter with just a for loop (no function)



clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
}

col_counter = 1

for key in clefairy:
    ws.cell(row=1, column=col_counter, value=clefairy[key])
    col_counter += 1

# create a function that takes in a pokemon

def write_to_xl(pokemon):
    col_


weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32
}

# call the function with weedle!
