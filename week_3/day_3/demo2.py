import json, requests
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Rick and Morty Locations")
ws.title = "Posts"
 

url = "https://jsonplaceholder.typicode.com/posts"
response = requests.get(url)
json_data = response.text
data = json.loads(json_data)

# print(data)

for row, post in enumerate(data, 1):
    for column, post_value in enumerate(post.values(), 1):
        ws.cell(row=row, column=column, value=str(post_value))
        ws1.cell(row=row, column=column, value=str(post_value))
wb.save("posts.xlsx")

