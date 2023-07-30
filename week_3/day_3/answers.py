# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# EASY MODE

# import the appropriate modules (you have 3)
import json, requests
from openpyxl import Workbook
character_url = "https://rickandmortyapi.com/api/character"

# set up a workbook and worksheet titled "Rick and Morty Characters"

wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

response = requests.get(character_url)
data = response.json()

characters = data['results']

column = 1

for key in characters[0].keys():
    ws.cell(row=1, column=column, value=key)
    column += 1

for row, character in enumerate(characters, 2):
    for column, character in enumerate(character.values(), 1):
        ws.cell(row=row, column=column, value=str(character))


ws2 = wb.create_sheet("Rick and Morty Locations")
ws3 = wb.create_sheet("Rick and Morty Episodes")

location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

response2 = requests.get(location_url)
data2 = response2.json()
locations = data2['results']

response3 = requests.get(episode_url)
data3 = response3.json()
episodes = data3['results']

column = 1
for key in locations[0].keys():
    ws2.cell(row=1, column=column, value=key)
    column += 1

for row, location in enumerate(locations, 2):
    for column, location in enumerate(location.values(), 1):
        ws2.cell(row=row, column=column, value=str(location))

column = 1
for key in episodes[0].keys():
    ws3.cell(row=1, column=column, value=key)
    column += 1

for row, episode in enumerate(episodes, 2):
    for column, episode in enumerate(episode.values(), 1):
        ws3.cell(row=row, column=column, value=str(episode))



column = 1
for key in locations[0].keys():
    ws2.cell(row=1, column=column, value=key)
    column += 1

column = 1
for key in episodes[0].keys():
    ws3.cell(row=1, column=column, value=key)
    column += 1

while data2['info']['next']:
    url = data2['info']['next']
    response2 = requests.get(url)
    data2 = response2.json()
    locations += data2['results']

for row, location in enumerate(locations, 2):
    for column, location in enumerate(location.values(), 1):
        ws2.cell(row=row, column=column, value=str(location))


while data3['info']['next']:
    url = data3['info']['next']
    response3 = requests.get(url)
    data3 = response3.json()
    episodes += data3['results']

for row, episode in enumerate(episodes, 2):
    for column, episode in enumerate(episode.values(), 1):
        ws3.cell(row=row, column=column, value=str(episode))


wb.save("rick_and_morty_answers.xlsx")

#Nightmare
url = "https://rickandmortyapi.com/api/character"
response = requests.get(url)
data = response.json()
characters = data['results']

ws['A1'] = "Name"
ws['B1'] = "Status"
ws['C1'] = "Species"
ws['D1'] = "Gender"
ws['E1'] = "Location"
ws['F1'] = "Origin"
ws['G1'] = "Episodes"

while data['info']['next']:
    url = data['info']['next']
    response = requests.get(url)
    data = response.json()
    characters += data['results']

for row, character in enumerate(characters, 2):
    try:
        location_url = character['location']['url']
        location_response = requests.get(location_url)
        location_data = location_response.json()
        location_name = location_data['name']
    except:
        location_name = "Unknown"
    try:
        origin_url = character['origin']['url']
        origin_response = requests.get(origin_url)
        origin_data = origin_response.json()
        origin_name = origin_data['name']
    except:
        origin_name = "Unknown"

    episode_list = []
    for episode_url in character['episode']:
        try:
            episode_response = requests.get(episode_url)
            episode_data = episode_response.json()
            episode_name = episode_data['name']
            episode_list.append(episode_name)
        except:
            episode_name = "Unknown"
        episodes = '. '.join(episode_list)

    ws.cell(row=row, column=1, value=str('name'))
    ws.cell(row=row, column=2, value=str('status'))
    ws.cell(row=row, column=3, value=str('species'))
    ws.cell(row=row, column=4, value=str('gender'))
    ws.cell(row=row, column=5, value=str('location'))
    ws.cell(row=row, column=6, value=str('origin'))
    ws.cell(row=row, column=7, value=str('episodes'))



# wb.save("rick_and_morty_answers_nightmare.xlsx")


