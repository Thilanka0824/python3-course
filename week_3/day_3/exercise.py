# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)

character_url = "https://rickandmortyapi.com/api/character"

# EASY MODE

# import the appropriate modules (you have 3)
import json, requests
from openpyxl import Workbook

# set up a workbook and worksheet titled "Rick and Morty Characters"
wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"


# assign a variable 'data' with the returned GET request
response = requests.get(character_url)
json_data = response.text
data = json.loads(json_data)
# print(data)


# create the appropriate headers in openpyxl for all of the keys for a single character
headers = []
for key in data['results'][0].keys():
  headers.append(key)
ws.append(headers)


# wb.save("rick_and_morty.xlsx")

# print(headers)


# loop through all of the 'results' of the data to populate the rows and columns for each character
for row, character in enumerate(data['results'], 2):
  for column, char_value in enumerate(character.values(), 1):
    ws.cell(row=row, column=column, value=str(char_value))

# due to the headers, the rows need to be offset by one!

wb.save("rick_and_morty.xlsx")
# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"
ws1 = wb.create_sheet("Rick and Morty Locations")
ws2 = wb.create_sheet("Rick and Morty Episodes")

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
episode_url = "https://rickandmortyapi.com/api/episode"
location_url = "https://rickandmortyapi.com/api/location"

# populate the new worksheets appropriately with all of the data!
response2 = requests.get(location_url)
data2 = response2.json()
locations = data2['results']

response3 = requests.get(episode_url)
data3 = response3.json()
episodes = data3['results']

#episodes
epi_response = requests.get(episode_url)
epi_json_data = epi_response.text
epi_data = json.loads(epi_json_data)

epi_headers = []
for key in epi_data['results'][0].keys():
  epi_headers.append(key)
ws2.append(epi_headers)

for row, episode in enumerate(epi_data["results"], 2):
  for column, episode_value in enumerate(episode.values(), 1):
    ws2.cell(row=row, column=column, value=str(episode_value))
wb.save("rick_and_morty.xlsx")

#location
loc_response = requests.get(location_url)
loc_json_data = loc_response.text
loc_data = json.loads(loc_json_data)

loc_headers = []
for key in loc_data['results'][0].keys():
  loc_headers.append(key)
ws1.append(loc_headers)

for row, location in enumerate(loc_data["results"], 2):
  for column, location_value in enumerate(location.values(), 1):
    ws1.cell(row=row, column=column, value=str(location_value))
wb.save("rick_and_morty.xlsx")


# use the "next" URL to continuously pull data until all data has been retrieved
# while url is not None:
#   response = requests.get(url)
#   json_data = response.text
#   data = json.loads(json_data)
  
#   # add the new data to the worksheet
#   for row, character in enumerate(data['results'], len(ws['A'])+1):
#     for column, char_value in enumerate(character.values(), 1):
#       ws.cell(row=row, column=column, value=str(char_value))
  
#   # update the "next" URL
#   url = data['info'].get('next')



# NOTE: don't forget your headers!

# HARD MODE

# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)
ws2 = wb.create_sheet("Rick and Morty Locations")
ws3 = wb.create_sheet("Rick and Morty Episodes")

column = 1
for key in locations[0].keys():
    ws2.cell(row=1, column=column, value=key)
    column += 1

column = 1
for key in episodes[0].keys():
    ws3.cell(row=1, column=column, value=key)
    column += 1

while data2['info']['next']:
    url = data2['info']['next']
    response2 = requests.get(url)
    data2 = response2.json()
    locations += data2['results']

for row, location in enumerate(locations, 2):
    for column, location in enumerate(location.values(), 1):
        ws2.cell(row=row, column=column, value=str(location))


while data3['info']['next']:
    url = data3['info']['next']
    response3 = requests.get(url)
    data3 = response3.json()
    episodes += data3['results']

for row, episode in enumerate(episodes, 2):
    for column, episode in enumerate(episode.values(), 1):
        ws3.cell(row=row, column=column, value=str(episode))

wb.save("rick_and_morty_answers.xlsx")
# NIGHTMARE

# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url

"""
{
  "id": 28,
  "name": "The Ricklantis Mixup",
  "air_date": "September 10, 2017",
  "episode": "S03E07",
  "characters": [
    "https://rickandmortyapi.com/api/character/1",
    "https://rickandmortyapi.com/api/character/2",
    ...
  ],
  "url": "https://rickandmortyapi.com/api/episode/28",
  "created": "2017-11-10T12:56:36.618Z"
}
"""

# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)

# for row, episode in enumerate(data, 2):
#   for column, episode in enumerate(episode.items(), 1):
#       if episode[0] == "characters":
#           # new list to collect the names
#           # for loop for each character  
#           # then do another request
#           # deserialize for just the name       
#           # append name to new list
#           ws.cell(row=row, column=column, value=str(name_list))
#       else:
#         ws.cell(row=row, column=column, value=str(episode[1]))


  # set the URL to the "next" URL in the response, if it exists