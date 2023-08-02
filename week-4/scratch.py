#imports
import json, requests
from openpyxl import load_workbook, Workbook

#loading the workbook
wb = load_workbook("/Users/ganegodagerodrigo/Desktop/VetsinTech/python3-course/week-4/project_day_1 copy.xlsx")

#referencing the worksheet
ws = wb['alpha_roster']
wt = wb["list_of_non_compliant_personnel"]
wt["A1"] = "emails of non compliant personnel"

#setting urls
basic_info_url = 'https://vit-py-api.herokuapp.com/employee_basic_info'
account_info_url = 'https://vit-py-api.herokuapp.com/employee_account_info'

#setting response data 
response = requests.get(basic_info_url)
json_data = response.text
data = json.loads(json_data)
basic_info = data['body']

#setting response data 
response =requests.get(account_info_url)
json_data = response.text
data = json.loads(json_data)
account_info = data['body']

#parse up username
def name_parser(username):
    first_name = ""
    last_name = ""
    new_string = ''.join((x for x in username if not x.isdigit()))
    if "." in new_string:
        split_un = new_string.split(".")
        first_name = split_un[0]
        last_name = split_un[1]
    elif "_" in new_string:
        split_un = new_string.split("_")
        first_name = split_un[0]    
        last_name = split_un[1]
    else: 
        first_name = new_string
        last_name = ""
    
    return {"first_name": first_name, "last_name":last_name}

#write non compliant users to other tab in workbook
def non_compliant_users(employee, index):
    wt.cell(row = index, column = 1, value = employee)
# give other tab a heading
ws["E1"] = "emails"

#Where the fun begins, writing api info to alpha_roster tab
for index, employee in enumerate(basic_info, start=0):
    names = name_parser(employee["username"])
    ws.cell(row=index+2, column=1, value=employee["userId"])
    ws.cell(row=index+2, column=2, value=names["first_name"])
    ws.cell(row=index+2, column=3, value=names["last_name"])
    ws.cell(row=index+2, column=4, value=account_info[index]["daysSinceLastPasswordChange"])
    ws.cell(row=index+2, column=5, value=account_info[index]["userEmail"])
    #writing non compliant users to non compliant tab
    if int(account_info[index]["daysSinceLastPasswordChange"]) > 60:
        non_compliant_users(account_info[index]["userEmail"], (index+2))

#getting rid of white space in non compliant tab 
colA = wt["A"]
clean_email_list = []
#getting rid of white space in non compliant tab 
for cell in colA:
    if cell.value:
        clean_email_list.append(cell.value)
    else:
        delete_index = colA.index(cell) + 1
        print(delete_index)
#getting rid of white space in non compliant tab 
wt.delete_cols(1)
for index, value in enumerate(clean_email_list, 1):
    wt.cell(row = index, column = 1, value = clean_email_list[index-1])

#saving work book
wb.save("employee_sheet.xlsx")
